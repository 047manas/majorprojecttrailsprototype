from app.models import db, User, StudentActivity, ActivityType
from sqlalchemy import func, case, or_, and_, distinct, extract, Integer, tuple_, literal
from flask_login import current_user
from flask import url_for
from datetime import datetime
import pandas as pd
import io
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

class AnalyticsService:

# ... (Previous code remains unchanged until _get_event_summary_list) ...

    @staticmethod
    def _get_event_summary_list(filters=None):
        """
        Helper for Sheet 2: Event Summary
        Groups strictly by Identity.
        """
        base_q = AnalyticsService._get_base_query(filters)
        base_q = base_q.outerjoin(ActivityType, StudentActivity.activity_type_id == ActivityType.id)
        
        event_date_expr = AnalyticsService._get_event_date_expr()
        identity_expr = AnalyticsService._get_event_identity_expr()
        
        q = base_q.with_entities(
            func.coalesce(ActivityType.name, StudentActivity.custom_category).label('category'),
            # For Title: If Defined, ignore title (use ''), else use actual title
            case((ActivityType.id.isnot(None), literal('')), else_=func.lower(func.trim(StudentActivity.title))).label('title_key'),
            func.max(StudentActivity.title).label('raw_title'), # FIXED: Use Aggregate to avoid GroupingError
            event_date_expr.label('date'),
            func.count(StudentActivity.id).label('participations'),
            func.count(distinct(StudentActivity.student_id)).label('unique_students'),
             func.sum(case((or_(StudentActivity.status == 'faculty_verified', StudentActivity.status == 'auto_verified'), 1), else_=0)).label('verified_count')
        ).group_by(
            identity_expr,
            ActivityType.id, # Explicitly group by ID to allow Name selection
            ActivityType.name,
            StudentActivity.custom_category,
            event_date_expr,
            case((ActivityType.id.isnot(None), literal('')), else_=func.lower(func.trim(StudentActivity.title)))
        )
        
        results = q.all()
        
        # Formatting for Excel
        # Title display: If category matches key (defined), show Category as Title or Generic.
        data = []
        for r in results:
            # If title_key is empty, it was a defined event. We can use Category name as title or leave blank.
            display_title = r.category if r.title_key == '' else r.raw_title
            
            data.append({
                "Event Category": r.category,
                "Event Title": display_title, # Fallback
                "Event Date": r.date,
                "Total Participants": r.participations,
                "Unique Students": r.unique_students,
                "Verified Count": int(r.verified_count or 0),
                "Engagement %": f"{round(r.unique_students/r.participations*100, 1) if r.participations else 0}%"
            })
        return data

    @staticmethod
    def get_comparative_stats(filters=None):
        """
        [NEW] Comparative Year-on-Year Growth
        """
        if not filters or not filters.get('year'):
            return None 
            
        import time
        start_time = time.time()
        
        try:
            current_year = int(filters['year'])
        except:
            return None
            
        # 1. Current Year Stats
        current_stats = AnalyticsService.get_institution_kpis(filters)
        
        # 2. Previous Year Stats
        prev_filters = filters.copy()
        prev_filters['year'] = current_year - 1
        prev_stats = AnalyticsService.get_institution_kpis(prev_filters)
        
        # 3. Calculate Growth
        comparison = {}
        for key, val in current_stats.items():
            prev_val = prev_stats.get(key, 0)
            
            growth_pct = None
            label = "No Data"
            
            if prev_val > 0:
                diff = val - prev_val
                growth_pct = round((diff / prev_val) * 100, 1)
                label = f"{growth_pct}%"
            elif prev_val == 0 and val > 0:
                label = "New"
            elif prev_val == 0 and val == 0:
                label = "-"
            
            comparison[key] = {
                "current": val,
                "previous": prev_val,
                "growth_pct": growth_pct,
                "label": label
            }
            
        elapsed = (time.time() - start_time) * 1000
        print(f"COMPARISON EXEC TIME: {elapsed:.2f}ms")
        
        return comparison

    @staticmethod
    def generate_naac_excel(filters=None, export_type='full'):
        """
        [REFACTORED] 6-Sheet Audit Ready Report
        """
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        
        # --- SHEET 1: Executive Summary ---
        if export_type in ['full']:
            # 1. KPIs
            kpis = AnalyticsService.get_institution_kpis(filters)
            # 2. Insights
            insights = AnalyticsService.get_admin_insights(filters)
            # 3. Growth
            growth = AnalyticsService.get_comparative_stats(filters)
            
            summary_data = {
                "Report Date": datetime.now().strftime("%Y-%m-%d"),
                "Applied Filters": str(filters),
                "Total Students": kpis['total_students'],
                "Engagement Rate": f"{kpis['engagement_rate']}%",
                "Verification Rate": f"{kpis['verified_rate']}%",
                "Top Department": f"{insights['top_dept']} ({insights['top_dept_val']}%)",
                "Top Event": f"{insights['top_event']} ({insights['top_event_val']})",
                "Risk Events Count": len(insights['risk_events']),
                "Verification Efficiency": f"{insights['verification_efficiency']}%"
            }
            
            if growth:
                summary_data["YoY Growth (Engagement)"] = growth['engagement_rate']['label']
                summary_data["YoY Growth (Events)"] = growth['total_events']['label']

            df1 = pd.DataFrame([summary_data])
            df1.to_excel(writer, sheet_name='Executive_Summary', index=False)
            AnalyticsService._format_excel_sheet(writer, df1, 'Executive_Summary')

        # --- SHEET 2: Event Intelligence ---
        if export_type in ['full', 'events']:
            events = AnalyticsService._get_event_summary_list(filters)
            # Add Risk Flag
            for e in events:
                total = e['Total Participants']
                pending = e.get('Pending Count', 0)
                e['Risk Flag'] = "HIGH PENDING" if (total > 0 and (pending/total) > 0.4) else ""
            
            df2 = pd.DataFrame(events)
            df2.to_excel(writer, sheet_name='Event_Intelligence', index=False)
            AnalyticsService._format_excel_sheet(writer, df2, 'Event_Intelligence')

        # --- SHEET 3: Department Performance ---
        if export_type in ['full']:
            dept_stats = AnalyticsService.get_department_participation(filters)
            if isinstance(dept_stats, dict): dept_stats = []
            
            # Ranking included implicity by sort order, but let's add explicitly
            for i, d in enumerate(dept_stats):
                d['Rank'] = i + 1
                
            s3_data = [{
                "Rank": d['Rank'],
                "Department": d['department'],
                "Total Students": d['total'],
                "Participated": d.get('unique', 0),
                "Engagement %": f"{d['engagement_percent']}%",
                "Events": d['events']
            } for d in dept_stats]
            
            df3 = pd.DataFrame(s3_data)
            df3.to_excel(writer, sheet_name='Dept_Performance', index=False)
            AnalyticsService._format_excel_sheet(writer, df3, 'Dept_Performance')

        # --- SHEET 4: Student Participation ---
        if export_type in ['full', 'students']:
            # Reuse logic - filtered
            raw_rows = AnalyticsService.get_student_list(paginate=False, filters=filters)
            s4_data = []
            for item in raw_rows:
                 # Certificate Link Logic
                cert_link = "Not Available"
                if item.certificate_file:
                    try:
                         cert_link = url_for('student.serve_upload', filename=item.certificate_file, _external=True)
                    except: cert_link = item.certificate_file

                s4_data.append({
                    "Student": item.student.full_name,
                    "Roll No": item.student.institution_id,
                    "Dept": item.student.department,
                    "Event": item.title,
                    "Category": item.activity_type.name if item.activity_type else (item.custom_category or 'Other'),
                    "Date": str(item.start_date or item.created_at.date()),
                    "Status": item.status,
                    "Certificate": cert_link
                })
            df4 = pd.DataFrame(s4_data)
            df4.to_excel(writer, sheet_name='Student_Participation', index=False)
            AnalyticsService._format_excel_sheet(writer, df4, 'Student_Participation')

        # --- SHEET 5: Audit Flags (Full Export Only) ---
        if export_type == 'full':
            health = AnalyticsService.get_data_health_summary()
            s5_data = [{
                "Metric": k, "Value": v
            } for k, v in health.items()]
            df5 = pd.DataFrame(s5_data)
            df5.to_excel(writer, sheet_name='Audit_Flags', index=False)
            AnalyticsService._format_excel_sheet(writer, df5, 'Audit_Flags')

        # --- SHEET 6: Yearly Growth (Full Only) ---
        if export_type == 'full':
            try:
                trend_filters = filters.copy() if filters else {}
                if 'year' in trend_filters: del trend_filters['year']
                
                trend_data = AnalyticsService.get_yearly_trend(trend_filters)
                if isinstance(trend_data, dict): trend_data = []
                
                df6 = pd.DataFrame(trend_data)
                df6.to_excel(writer, sheet_name='Yearly_Growth', index=False)
                AnalyticsService._format_excel_sheet(writer, df6, 'Yearly_Growth')
            except Exception as e:
                print(f"Error generating Yearly Growth sheet: {e}")

        writer.close()
        output.seek(0)
        return output

# ... (rest of class)

    @staticmethod
    def _format_excel_sheet(writer, df, sheet_name):
        """
        Helper: Apply formatting (Bold Header, Auto Width)
        """
        try:
            worksheet = writer.sheets[sheet_name]
            
            # Check if dataframe is empty
            if df.empty:
                return

            for idx, col in enumerate(df.columns):
                # Header Bold
                cell = worksheet.cell(row=1, column=idx+1)
                cell.font = Font(bold=True)
                
                # Auto Width (Basic Approximation)
                # Handle empty column data safety
                col_data = df[col].astype(str)
                max_len = 0
                if not col_data.empty:
                     max_len = col_data.map(len).max()
                
                final_width = max((max_len, len(str(col)))) + 2
                worksheet.column_dimensions[get_column_letter(idx+1)].width = min(final_width, 50)
        except Exception as e:
            print(f"Excel Formatting Error on {sheet_name}: {e}")


    @staticmethod
    def _apply_role_scope(query):
        """
        [PART 4] STRICT ROLE-BASED DATA SCOPE
        """
        from flask import has_request_context
        if not has_request_context():
            return query  # CLI/test context: no scope restriction
        if not current_user or not current_user.is_authenticated:
            return query.filter(1 == 0) # No access
            
        if current_user.role == 'admin':
            return query
            
        if current_user.role == 'faculty':
            conditions = []
            
            # 1. HOD: Filter by User.department
            if current_user.position and current_user.position.lower() == 'hod':
                if current_user.department:
                    conditions.append(User.department == current_user.department)
            
            # 2. Event In-Charge: Filter by Activity Types managed
            managed_activities = ActivityType.query.filter_by(faculty_incharge_id=current_user.id).all()
            if managed_activities:
                managed_ids = [a.id for a in managed_activities]
                conditions.append(StudentActivity.activity_type_id.in_(managed_ids))
                
            if conditions:
                return query.filter(or_(*conditions))
            else:
                 return query.filter(1 == 0)

        if current_user.role == 'student':
             return query.filter(StudentActivity.student_id == current_user.id)
             
        return query.filter(1 == 0)

    @staticmethod
    def _get_event_date_expr():
        """
        Helper: Coalesce start_date or created_at (cast to date)
        MANDATORY: Use this everywhere for date logic.
        """
        return func.coalesce(StudentActivity.start_date, func.cast(StudentActivity.created_at, db.Date))

    @staticmethod
    def _get_event_identity_expr():
        """
        Helper: Unique Event Identity - STRICT RULE
        - IF Activity Type is Defined (ID Not None): 'TYPE-' + ID + '-' + Date
        - ELSE (Custom): 'CUSTOM-' + Clean Title + '-' + Date
        """
        # Logic: Case when ActivityTypeID is NOT NULL -> Use TYPE-ID-Date
        # Logic: Case when ActivityTypeID IS NULL -> Use CUSTOM-Title-Date
        
        event_date = str(func.coalesce(StudentActivity.start_date, literal('nodate')))
        
        return case(
            (StudentActivity.activity_type_id.isnot(None),
             func.concat('TYPE-', StudentActivity.activity_type_id, '-', func.coalesce(func.cast(StudentActivity.start_date, db.String), 'nodate'))),
            else_=func.concat('CUSTOM-', func.coalesce(StudentActivity.custom_category, 'other'), '-', func.lower(func.trim(StudentActivity.title)), '-', func.coalesce(func.cast(StudentActivity.start_date, db.String), 'nodate'))
        )

    @staticmethod
    def _apply_filters(query, filters):
        """
        Dynamic Query Builder - Stateless
        """
        if not filters:
            return query
        
        # Use coalesced date for filtering
        event_date = AnalyticsService._get_event_date_expr()
            
        # 1. Academic Year
        if filters.get('year'):
            try:
                year = int(filters['year'])
                query = query.filter(extract('year', event_date) == year)
            except: pass

        # 2. Department
        if filters.get('department'):
            query = query.filter(User.department == filters['department'])
            
        # 3. Batch
        if filters.get('batch'):
             query = query.filter(User.batch_year == str(filters['batch']))

        # 4. Verified Only
        if filters.get('verified_only'):
             query = query.filter(or_(
                StudentActivity.status == 'faculty_verified',
                StudentActivity.status == 'auto_verified'
            ))
            
        # 5. Date Range
        if filters.get('start_date'):
            try:
                sd = filters['start_date']
                query = query.filter(event_date >= sd)
            except: pass
            
        if filters.get('end_date'):
            try:
                ed = filters['end_date']
                query = query.filter(event_date <= ed)
            except: pass
            
        # 6. Activity Type
        if filters.get('activity_type_id') or filters.get('event_type_id'):
            tid = filters.get('activity_type_id') or filters.get('event_type_id')
            try:
                query = query.filter(StudentActivity.activity_type_id == int(tid))
            except: pass

        # 7. Specific Event Identity (For Drilldown)
        if filters.get('event_identity'):
            identity_expr = AnalyticsService._get_event_identity_expr()
            query = query.filter(identity_expr == filters['event_identity'])
            
        return query

    @staticmethod
    def _get_base_query(filters=None):
        """
        SINGLE SOURCE OF TRUTH
        All data retrieval must start here.
        """
        query = db.session.query(StudentActivity).join(User, StudentActivity.student_id == User.id)
        query = AnalyticsService._apply_role_scope(query)
        query = AnalyticsService._apply_filters(query, filters)
        return query

    @staticmethod
    def get_test_student_list(activity_type_id):
        return []

    @staticmethod
    def get_institution_kpis(filters=None):
        """
        [FIXED] KPIs with Strict Identity & Zero State
        """
        base_q = AnalyticsService._get_base_query(filters)
        
        # 1. Total Students (Active) - Sourced from User table directly for normalization
        # Note: 'Total Students' in KPI usually means relevant students context. 
        # But per request, let's keep it scoped to Active students in DB matching Dept/Batch filters.
        total_students_q = db.session.query(func.count(User.id)).filter(User.role == 'student', User.is_active == True)
        if filters:
            if filters.get('department'):
                total_students_q = total_students_q.filter(User.department == filters['department'])
            if filters.get('batch'):
                total_students_q = total_students_q.filter(User.batch_year == str(filters['batch']))
        total_students = total_students_q.scalar() or 0

        # 2. Total Events - Strict Identity
        total_events = base_q.with_entities(
            func.count(distinct(AnalyticsService._get_event_identity_expr()))
        ).scalar() or 0
        
        # 3. Total Participations
        total_participations = base_q.with_entities(func.count(StudentActivity.id)).scalar() or 0
        
        # 4. Unique Students
        unique_students = base_q.with_entities(func.count(distinct(StudentActivity.student_id))).scalar() or 0
        
        # 5. Engagement Rate
        engagement_rate = round((unique_students / total_students * 100), 1) if total_students > 0 else 0
        
        # 6. Avg Activities
        avg_activities_per_student = round((total_participations / unique_students), 2) if unique_students > 0 else 0
        
        # 7. Verified Rate
        verified_count = base_q.filter(or_(
            StudentActivity.status == 'faculty_verified',
            StudentActivity.status == 'auto_verified'
        )).with_entities(func.count(StudentActivity.id)).scalar() or 0
        
        verified_rate = round((verified_count / total_participations * 100), 1) if total_participations > 0 else 0

        print(f"DEBUG KPI: Events={total_events}, Part={total_participations}, Unique={unique_students}")

        return {
            "total_students": total_students,
            "total_events": total_events,
            "total_participations": total_participations,
            "unique_students": unique_students,
            "engagement_rate": engagement_rate,
            "avg_activities_per_student": avg_activities_per_student,
            "verified_rate": verified_rate
        }

    @staticmethod
    def get_event_distribution(filters=None):
        """
        [FIXED] Group by Category with Strict Identity
        """
        base_q = AnalyticsService._get_base_query(filters)
        base_q = base_q.outerjoin(ActivityType, StudentActivity.activity_type_id == ActivityType.id)
        
        cat_name = func.coalesce(ActivityType.name, 'Other / Custom').label('category_name')
        
        query = base_q.with_entities(
            cat_name,
            func.count(distinct(AnalyticsService._get_event_identity_expr())).label('total_events'),
            func.count(StudentActivity.id).label('participations')
        ).group_by(cat_name)
        
        results = query.all()
        if not results:
            return {"empty": True}
            
        return [{
            "category": r.category_name, 
            "count": r.total_events,
            "participations": r.participations
        } for r in results]

    @staticmethod
    def get_department_participation(filters=None):
        """
        [FIXED] Dept Participation
        """
        # Note: Must align with _get_base_query logic, but group by Dept.
        # Calling _get_base_query ensures filters are applied.
        base_q = AnalyticsService._get_base_query(filters)
        
        query = base_q.with_entities(
            User.department,
            func.count(distinct(StudentActivity.student_id)).label('participated_students'),
            func.count(distinct(AnalyticsService._get_event_identity_expr())).label('events'),
            func.count(StudentActivity.id).label('participations')
        ).group_by(User.department)
        
        results = query.all()
        if not results:
            return {"empty": True}

        # Fetch total students per dept for normalization (Active Only context)
        all_depts = db.session.query(User.department, func.count(User.id))\
            .filter(User.role == 'student', User.is_active == True)\
            .group_by(User.department).all()
        dept_counts = {r[0]: r[1] for r in all_depts if r[0]}
        
        data = []
        for r in results:
            if not r.department: continue
            total = dept_counts.get(r.department, 0)
            rate = round((r.participated_students / total * 100), 1) if total > 0 else 0
            
            data.append({
                "department": r.department,
                "engagement_percent": rate,
                "participated": r.participated_students,
                "events": r.events,
                "participations": r.participations,
                "unique": r.participated_students, # alias for excel
                "total": total
            })
            
        return sorted(data, key=lambda x: x['engagement_percent'], reverse=True)

    @staticmethod
    def get_yearly_trend(filters=None):
        """
        [FIXED] Yearly Trend with Null Handling (Year 0)
        """
        base_q = AnalyticsService._get_base_query(filters)
        
        event_date = AnalyticsService._get_event_date_expr()
        year_expr = extract('year', event_date).label('year')
        
        query = base_q.with_entities(
            year_expr,
            func.count(distinct(AnalyticsService._get_event_identity_expr())).label('total_events'),
            func.count(StudentActivity.id).label('total_participations')
        ).group_by(year_expr).order_by(year_expr)
        
        results = query.all()
        if not results:
            return {"empty": True}
            
        return [{
            "year": int(r.year) if r.year != 0 else "Unspecified",
            "total_events": r.total_events,
            "total_participations": r.total_participations
        } for r in results]

    @staticmethod
    def get_verification_summary(filters=None):
        """
        [FIXED] Verification Logic
        """
        base_q = AnalyticsService._get_base_query(filters)
        
        query = base_q.with_entities(
            func.sum(case((or_(StudentActivity.status == 'faculty_verified', StudentActivity.status == 'auto_verified'), 1), else_=0)),
            func.sum(case((StudentActivity.status == 'pending', 1), else_=0)),
            func.sum(case((StudentActivity.status == 'rejected', 1), else_=0))
        )
        
        row = query.first()
        if not row:
             return {"empty": True}
             
        verified = int(row[0] or 0)
        pending = int(row[1] or 0)
        rejected = int(row[2] or 0)
        
        if (verified + pending + rejected) == 0:
            return {"empty": True}

        return {
            "verified": verified,
            "not_verified": pending + rejected,
            "details": {"pending": pending, "rejected": rejected}
        }

    @staticmethod
    def _build_student_query(base_q, category_name=None, department=None, search=None, status=None):
        """
        Shared filter builder for student list queries.
        Reused by get_student_list and export helpers.
        """
        if department and department != 'All':
            base_q = base_q.filter(User.department == department)
        
        if status and status != 'All':
            base_q = base_q.filter(StudentActivity.status == status)

        if category_name and category_name != 'All':
            if category_name == 'Other / Custom':
                base_q = base_q.outerjoin(ActivityType, StudentActivity.activity_type_id == ActivityType.id)
                base_q = base_q.filter(or_(ActivityType.id.is_(None), ActivityType.name.is_(None)))
            else:
                base_q = base_q.join(ActivityType, StudentActivity.activity_type_id == ActivityType.id)
                base_q = base_q.filter(ActivityType.name == category_name)
        
        if search:
            search_term = f"%{search}%"
            base_q = base_q.filter(or_(
                User.full_name.ilike(search_term),
                User.institution_id.ilike(search_term),
                StudentActivity.title.ilike(search_term)
            ))
        
        return base_q

    @staticmethod
    def _serialize_student_item(item, include_certificate=False):
        """Serialize a StudentActivity ORM object to dict."""
        row = {
            "student_name": item.student.full_name,
            "roll_number": item.student.institution_id,
            "department": item.student.department,
            "title": item.title,
            "status": item.status,
            "category": item.activity_type.name if item.activity_type else (item.custom_category or 'Other'),
            "date": str(item.start_date or item.created_at.date()),
            "verification_mode": item.verification_mode or 'N/A'
        }
        if include_certificate:
            cert_url = None
            if item.certificate_file:
                try:
                    cert_url = url_for('student.serve_upload', filename=item.certificate_file, _external=True)
                except Exception:
                    cert_url = None
            row["certificate_url"] = cert_url
            row["certificate_hash"] = item.certificate_hash
            row["verification_token"] = item.verification_token
        return row

    @staticmethod
    def get_student_list(category_name=None, department=None, page=1, per_page=20, filters=None, search=None, status=None, paginate=True):
        """
        Drilldown List — now includes certificate data for admin/faculty.
        """
        import time
        t0 = time.time()

        base_q = AnalyticsService._get_base_query(filters)
        base_q = AnalyticsService._build_student_query(base_q, category_name, department, search, status)

        query = base_q.order_by(AnalyticsService._get_event_date_expr().desc())
        
        if not paginate:
            return query.all()
        
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        # Determine if caller has certificate access
        include_cert = False
        try:
            from flask import has_request_context
            if has_request_context() and current_user and current_user.is_authenticated:
                include_cert = current_user.role in ('admin', 'faculty')
        except Exception:
            pass

        elapsed = (time.time() - t0) * 1000
        if elapsed > 1000:
            print(f"PERF WARNING: get_student_list took {elapsed:.0f}ms")
        
        return {
            "students": [AnalyticsService._serialize_student_item(item, include_certificate=include_cert) for item in pagination.items],
            "total_pages": pagination.pages,
            "current_page": pagination.page,
            "total_records": pagination.total
        }

    @staticmethod
    def _get_event_summary_list(filters=None):
        """
        Helper for Sheet 2: Event Summary
        Groups strictly by Identity.
        """
        base_q = AnalyticsService._get_base_query(filters)
        base_q = base_q.outerjoin(ActivityType, StudentActivity.activity_type_id == ActivityType.id)
        
        event_date_expr = AnalyticsService._get_event_date_expr()
        identity_expr = AnalyticsService._get_event_identity_expr()
        
        q = base_q.with_entities(
            # Aggregate non-grouped columns to satisfy SQL Grouping rules
            func.max(func.coalesce(ActivityType.name, StudentActivity.custom_category)).label('category'),
            func.max(case((ActivityType.id.isnot(None), literal('')), else_=func.lower(func.trim(StudentActivity.title)))).label('title_key'),
            func.max(StudentActivity.title).label('raw_title'),
            func.max(event_date_expr).label('date'), # Date is part of identity, so max(date) == date
            
            func.count(StudentActivity.id).label('participations'),
            func.count(distinct(StudentActivity.student_id)).label('unique_students'),
             func.sum(case((or_(StudentActivity.status == 'faculty_verified', StudentActivity.status == 'auto_verified'), 1), else_=0)).label('verified_count'),
             func.sum(case((StudentActivity.status == 'pending', 1), else_=0)).label('pending_count')
        ).group_by(
            identity_expr
        )
        
        results = q.all()
        
        # Formatting for Excel
        data = []
        for r in results:
            display_title = r.category if r.title_key == '' else r.raw_title
            
            data.append({
                "Event Category": r.category,
                "Event Title": display_title, 
                "Event Date": r.date,
                "Total Participants": r.participations,
                "Unique Students": r.unique_students,
                "Verified Count": int(r.verified_count or 0),
                "Pending Count": int(r.pending_count or 0),
                "Engagement %": f"{round(r.unique_students/r.participations*100, 1) if r.participations else 0}%"
            })
        return data

    @staticmethod
    def get_data_health_summary():
        """
        [NEW] Data Integrity & Health Check
        """
        base_q = AnalyticsService._get_base_query() # No filters for global health
        
        # 1. Null Dates
        null_dates = base_q.filter(StudentActivity.start_date.is_(None)).count()
        total_records = base_q.count()
        
        # 2. Missing Department
        missing_dept = base_q.filter(User.department.is_(None)).count()
        
        # 3. Duplicate Event Entries (Same Student, Same Identity)
        identity_expr = AnalyticsService._get_event_identity_expr()
        dupe_entries = base_q.with_entities(StudentActivity.student_id, identity_expr)\
            .group_by(StudentActivity.student_id, identity_expr)\
            .having(func.count(StudentActivity.id) > 1).count()

        # 4. Events Missing Category
        missing_cat = base_q.filter(
            StudentActivity.activity_type_id.is_(None),
            or_(StudentActivity.custom_category.is_(None), StudentActivity.custom_category == '')
        ).count()
        
        return {
            "total_records": total_records,
            "null_dates": null_dates,
            "null_dates_percent": round(null_dates/total_records*100, 1) if total_records else 0,
            "missing_dept": missing_dept,
            "missing_dept_percent": round(missing_dept/total_records*100, 1) if total_records else 0,
            "duplicate_entries": dupe_entries,
            "missing_category": missing_cat
        }

    @staticmethod
    def get_admin_insights(filters=None):
        """
        [NEW] Administrative Insights 
        Reuses existing aggregation methods.
        """
        import time
        start_time = time.time()
        
        insights = {
            "top_dept": "N/A", "top_dept_val": 0,
            "low_dept": "N/A", "low_dept_val": 0,
            "top_event": "N/A", "top_event_val": 0,
            "top_category": "N/A",  "top_category_val": 0,
            "verification_efficiency": 0,
            "low_engagement_depts": [],
            "risk_events": []
        }
        
        # 1. Dept Performance
        dept_stats = AnalyticsService.get_department_participation(filters)
        if dept_stats and not isinstance(dept_stats, dict):
            top = dept_stats[0]
            low = dept_stats[-1]
            insights['top_dept'] = top['department']
            insights['top_dept_val'] = top['engagement_percent']
            insights['low_dept'] = low['department']
            insights['low_dept_val'] = low['engagement_percent']
            
            insights['low_engagement_depts'] = [d['department'] for d in dept_stats if d['engagement_percent'] < 30]

        # 2. Event Performance
        events = AnalyticsService._get_event_summary_list(filters)
        if events:
            top_event = max(events, key=lambda x: x['Unique Students'])
            insights['top_event'] = top_event['Event Title']
            insights['top_event_val'] = top_event['Unique Students']
            
            risk_list = []
            for e in events:
                total = e['Total Participants']
                pending = e.get('Pending Count', 0)
                if total > 0 and (pending / total) > 0.4:
                    risk_list.append(e['Event Title'])
            insights['risk_events'] = risk_list

        # 3. Category Performance
        dist = AnalyticsService.get_event_distribution(filters)
        if dist and not isinstance(dist, dict):
            top_cat = max(dist, key=lambda x: x['participations'])
            insights['top_category'] = top_cat['category']
            insights['top_category_val'] = top_cat['participations']
            
        # 4. Verification Efficiency
        kpis = AnalyticsService.get_institution_kpis(filters)
        insights['verification_efficiency'] = kpis['verified_rate']

        elapsed = (time.time() - start_time) * 1000
        print(f"INSIGHTS EXEC TIME: {elapsed:.2f}ms")
        
        return insights

    @staticmethod
    def generate_naac_excel(filters=None, export_type='full'):
        """
        [FIXED] 4 Clean Sheets using Pandas & Service Reusability
        """
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        
        # --- SHEET 1: Institutional Summary ---
        if export_type in ['full']:
            kpis = AnalyticsService.get_institution_kpis(filters)
            s1_data = [{
                "Report Date": datetime.now().strftime("%Y-%m-%d"),
                "Applied Filters": str(filters),
                "Total Students (Active)": kpis['total_students'],
                "Total Events": kpis['total_events'],
                "Total Participations": kpis['total_participations'],
                "Unique Students": kpis['unique_students'],
                "Engagement Rate": f"{kpis['engagement_rate']}%",
                "Verified Rate": f"{kpis['verified_rate']}%",
                "Avg Activities/Student": kpis['avg_activities_per_student']
            }]
            df1 = pd.DataFrame(s1_data)
            df1.to_excel(writer, sheet_name='Institutional_Summary', index=False)
            AnalyticsService._format_excel_sheet(writer, df1, 'Institutional_Summary')

        # --- SHEET 2: Event Summary ---
        if export_type in ['full', 'events']:
            s2_data = AnalyticsService._get_event_summary_list(filters)
            df2 = pd.DataFrame(s2_data)
            df2.to_excel(writer, sheet_name='Event_Summary', index=False)
            AnalyticsService._format_excel_sheet(writer, df2, 'Event_Summary')

        # --- SHEET 3: Department Summary ---
        if export_type in ['full']:
            dept_stats = AnalyticsService.get_department_participation(filters)
            if isinstance(dept_stats, dict) and dept_stats.get('empty'):
                 dept_stats = []
            
            s3_data = [{
                "Department": d['department'],
                "Total Students": d['total'],
                "Participated Students": d.get('unique', 0),
                "Engagement %": f"{d['engagement_percent']}%",
                "Total Events": d['events'],
                "Total Participations": d['participations']
            } for d in dept_stats]
            
            df3 = pd.DataFrame(s3_data)
            df3.to_excel(writer, sheet_name='Department_Summary', index=False)
            AnalyticsService._format_excel_sheet(writer, df3, 'Department_Summary')
        
        # --- SHEET 4: Student Participation (Audit-Ready) ---
        if export_type in ['full', 'students']:
            raw_rows = AnalyticsService.get_student_list(paginate=False, filters=filters)
            
            s4_data = []
            for item in raw_rows:
                cert_link = "Not Available"
                if item.certificate_file:
                    try:
                        cert_link = url_for('student.serve_upload', filename=item.certificate_file, _external=True)
                    except Exception:
                        cert_link = item.certificate_file 

                s4_data.append({
                    "Student Name": item.student.full_name,
                    "Roll No": item.student.institution_id,
                    "Department": item.student.department,
                    "Batch": item.student.batch_year, 
                    "Activity Title": item.title,
                    "Category": item.activity_type.name if item.activity_type else (item.custom_category or 'Other'),
                    "Date": str(item.start_date or item.created_at.date()),
                    "Status": item.status,
                    "Verification Mode": item.verification_mode or 'N/A',
                    "Certificate Hash": item.certificate_hash or 'N/A',
                    "Certificate Link": cert_link
                })
            
            df4 = pd.DataFrame(s4_data)
            df4.to_excel(writer, sheet_name='Student_Participation', index=False)
            AnalyticsService._format_excel_sheet(writer, df4, 'Student_Participation')
        
        writer.close()
        output.seek(0)
        return output

    # ============================================================
    # PHASE 5: ADVANCED EXPORTS
    # ============================================================

    @staticmethod
    def get_comparative_stats(filters=None):
        """
        Year-over-Year comparison. Requires 'year' in filters.
        """
        if not filters or not filters.get('year'):
            return None
        
        current_year = int(filters['year'])
        prev_year = current_year - 1
        
        current_filters = {**filters, 'year': current_year}
        prev_filters = {**filters, 'year': prev_year}
        
        cur = AnalyticsService.get_institution_kpis(current_filters)
        prev = AnalyticsService.get_institution_kpis(prev_filters)
        
        def growth(cur_val, prev_val):
            if prev_val == 0:
                return {"current": cur_val, "previous": prev_val, "growth_pct": None}
            pct = round((cur_val - prev_val) / prev_val * 100, 1)
            return {"current": cur_val, "previous": prev_val, "growth_pct": pct}
        
        return {
            "current_year": current_year,
            "previous_year": prev_year,
            "total_events": growth(cur['total_events'], prev['total_events']),
            "total_participations": growth(cur['total_participations'], prev['total_participations']),
            "total_students": growth(cur['total_students'], prev['total_students']),
            "verified_rate": growth(cur['verified_rate'], prev['verified_rate'])
        }

    @staticmethod
    def generate_filtered_student_export(category_name=None, department=None, search=None, status=None, filters=None):
        """
        Export exactly the filtered student list as Excel.
        """
        import time
        t0 = time.time()

        base_q = AnalyticsService._get_base_query(filters)
        base_q = AnalyticsService._build_student_query(base_q, category_name, department, search, status)
        rows = base_q.order_by(AnalyticsService._get_event_date_expr().desc()).all()

        s_data = []
        for item in rows:
            cert_link = "Not Available"
            if item.certificate_file:
                try:
                    cert_link = url_for('student.serve_upload', filename=item.certificate_file, _external=True)
                except Exception:
                    cert_link = item.certificate_file
            
            s_data.append({
                "Student Name": item.student.full_name,
                "Roll No": item.student.institution_id,
                "Department": item.student.department,
                "Batch": item.student.batch_year,
                "Activity Title": item.title,
                "Category": item.activity_type.name if item.activity_type else (item.custom_category or 'Other'),
                "Date": str(item.start_date or item.created_at.date()),
                "Status": item.status,
                "Verification Mode": item.verification_mode or 'N/A',
                "Certificate Hash": item.certificate_hash or 'N/A',
                "Certificate Link": cert_link
            })

        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')

        # Summary row
        filter_desc = f"Dept={department or 'All'}, Cat={category_name or 'All'}, Status={status or 'All'}, Search={search or 'N/A'}"
        meta = pd.DataFrame([{
            "Generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "Records": len(s_data),
            "Filters Applied": filter_desc
        }])
        meta.to_excel(writer, sheet_name='Filtered_Student_List', index=False, startrow=0)

        df = pd.DataFrame(s_data)
        df.to_excel(writer, sheet_name='Filtered_Student_List', index=False, startrow=3)
        AnalyticsService._format_excel_sheet(writer, df, 'Filtered_Student_List')

        # Freeze top rows
        try:
            ws = writer.sheets['Filtered_Student_List']
            ws.freeze_panes = 'A5'
        except Exception:
            pass

        writer.close()
        output.seek(0)

        elapsed = (time.time() - t0) * 1000
        if elapsed > 1000:
            print(f"PERF WARNING: generate_filtered_student_export took {elapsed:.0f}ms")

        return output

    @staticmethod
    def generate_snapshot_export(filters=None):
        """
        Lightweight 3-sheet export: KPIs, Insights, Comparison.
        """
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')

        # Sheet 1: KPI Snapshot
        kpis = AnalyticsService.get_institution_kpis(filters)
        df1 = pd.DataFrame([{
            "Report Date": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "Filters": str(filters),
            "Total Students": kpis['total_students'],
            "Total Events": kpis['total_events'],
            "Participations": kpis['total_participations'],
            "Unique Students": kpis['unique_students'],
            "Engagement Rate": f"{kpis['engagement_rate']}%",
            "Verified Rate": f"{kpis['verified_rate']}%"
        }])
        df1.to_excel(writer, sheet_name='KPI_Snapshot', index=False)
        AnalyticsService._format_excel_sheet(writer, df1, 'KPI_Snapshot')

        # Sheet 2: Admin Insights
        insights = AnalyticsService.get_admin_insights(filters)
        df2 = pd.DataFrame([{
            "Top Department": insights['top_dept'],
            "Top Dept Engagement": f"{insights['top_dept_val']}%",
            "Lowest Department": insights.get('low_dept', 'N/A'),
            "Lowest Dept Engagement": f"{insights.get('low_dept_val', 0)}%",
            "Top Event": insights['top_event'],
            "Top Event Students": insights['top_event_val'],
            "Verification Efficiency": f"{insights['verification_efficiency']}%",
            "Risk Events Count": len(insights['risk_events']),
            "Risk Events": ', '.join(insights['risk_events']) if insights['risk_events'] else 'None'
        }])
        df2.to_excel(writer, sheet_name='Admin_Insights', index=False)
        AnalyticsService._format_excel_sheet(writer, df2, 'Admin_Insights')

        # Sheet 3: Comparison (if year available)
        comp = AnalyticsService.get_comparative_stats(filters)
        if comp:
            rows = []
            for metric in ['total_events', 'total_participations', 'total_students', 'verified_rate']:
                obj = comp[metric]
                rows.append({
                    "Metric": metric.replace('_', ' ').title(),
                    f"{comp['current_year']}": obj['current'],
                    f"{comp['previous_year']}": obj['previous'],
                    "Growth %": f"{obj['growth_pct']}%" if obj['growth_pct'] is not None else 'N/A'
                })
            df3 = pd.DataFrame(rows)
            df3.to_excel(writer, sheet_name='Year_Comparison', index=False)
            AnalyticsService._format_excel_sheet(writer, df3, 'Year_Comparison')
        else:
            df3 = pd.DataFrame([{"Note": "Select a specific Year filter to enable comparison."}])
            df3.to_excel(writer, sheet_name='Year_Comparison', index=False)

        writer.close()
        output.seek(0)
        return output

    @staticmethod
    def generate_event_instance_export(event_identity, filters=None):
        """
        Export students for a specific event identity (drilldown level).
        """
        base_q = AnalyticsService._get_base_query(filters)
        identity_expr = AnalyticsService._get_event_identity_expr()

        rows = base_q.filter(identity_expr == event_identity)\
            .order_by(User.department, User.full_name).all()

        s_data = []
        for item in rows:
            cert_link = "Not Available"
            if item.certificate_file:
                try:
                    cert_link = url_for('student.serve_upload', filename=item.certificate_file, _external=True)
                except Exception:
                    cert_link = item.certificate_file
            s_data.append({
                "Student Name": item.student.full_name,
                "Roll No": item.student.institution_id,
                "Department": item.student.department,
                "Batch": item.student.batch_year,
                "Status": item.status,
                "Verification Mode": item.verification_mode or 'N/A',
                "Certificate Hash": item.certificate_hash or 'N/A',
                "Certificate Link": cert_link
            })

        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')

        df = pd.DataFrame(s_data)
        df.to_excel(writer, sheet_name='Event_Instance_Report', index=False)
        AnalyticsService._format_excel_sheet(writer, df, 'Event_Instance_Report')

        writer.close()
        output.seek(0)
        return output
